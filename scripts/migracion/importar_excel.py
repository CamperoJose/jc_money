#!/usr/bin/env python3
"""
Migración del Excel My_Money_v5.0.xlsx -> Supabase (Postgres).

Fase 1, prioridad Patrimonio: migra la hoja CONTEOS a net_worth_snapshots +
net_worth_balances. (Gastos, DPF y Deudas se agregarán después.)

USO (desde tu PC, que sí llega a Supabase):

  1. pip install -r scripts/migracion/requirements.txt
  2. Aplica antes el esquema (0001) y las semillas (0002) en Supabase, y haz
     login al menos una vez en la app (para que exista tu usuario en auth.users).
  3. Asegúrate de tener DATABASE_URL en .env.local. Si la conexión directa falla
     (host solo-IPv6), usa la cadena del "Session pooler" (IPv4) que da Supabase
     en Project Settings -> Database -> Connection string.
  4. Dry-run (NO escribe, solo muestra qué haría):
        python scripts/migracion/importar_excel.py --excel ruta/My_Money_v5.0.xlsx
  5. Cargar de verdad:
        python scripts/migracion/importar_excel.py --excel ruta/My_Money_v5.0.xlsx --commit

Es idempotente por fecha: si ya existe una foto con la misma snapshot_date para
tu usuario, la salta (avisa).
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import sys

try:
    import openpyxl
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from dotenv import load_dotenv
except ImportError as e:
    sys.exit(f"Falta una dependencia: {e}. Corre: pip install -r scripts/migracion/requirements.txt")

# Encabezado (columna del Excel) -> nombre de la cuenta sembrada en la DB.
COLUMNA_A_CUENTA = {
    "Banco SOL": "Banco SOL",
    "Fortaleza": "Fortaleza",
    "BMSC": "BMSC",
    "BNB": "BNB",
    "IDEPRO CA": "IDEPRO CA",
    "EFECTIVO BS": "Efectivo Bs",
    "EFECTIVO USD": "Efectivo USD",
    "USDT": "USDT",
    "DPF CONGELADO": "DPF Congelado",
    "DEBTS": "Por Cobrar",  # la columna "Debts" es dinero por cobrar (activo)
}

HOJA_CONTEOS = "CONTEOS"
FILA_ENCABEZADO = 4  # los encabezados están en la fila 4


def cargar_env():
    # Carga .env.local si existe, sin pisar variables ya presentes.
    for archivo in (".env.local", ".env"):
        if os.path.exists(archivo):
            load_dotenv(archivo, override=False)


def conectar():
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("Falta DATABASE_URL (ponla en .env.local o expórtala).")
    try:
        return psycopg2.connect(url)
    except Exception as e:  # noqa: BLE001
        sys.exit(
            "No se pudo conectar a Postgres.\n"
            f"  {e}\n"
            "Si el host es solo-IPv6, usa la cadena del 'Session pooler' (IPv4) de Supabase."
        )


def obtener_user_id(cur) -> str:
    cur.execute("select id, email from auth.users order by created_at asc limit 1;")
    row = cur.fetchone()
    if not row:
        sys.exit("No hay usuarios en auth.users. Inicia sesión en la app una vez y reintenta.")
    print(f"  Usuario: {row['email']} ({row['id']})")
    return row["id"]


def obtener_cuentas(cur, user_id: str) -> dict[str, dict]:
    cur.execute(
        "select id, name, currency, is_liability from accounts where user_id = %s;",
        (user_id,),
    )
    cuentas = {r["name"]: r for r in cur.fetchall()}
    if not cuentas:
        sys.exit("No hay cuentas. Corre primero las semillas (0002_seed_catalogos.sql).")
    return cuentas


def leer_conteos(ruta_excel: str):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    if HOJA_CONTEOS not in wb.sheetnames:
        sys.exit(f"El Excel no tiene la hoja '{HOJA_CONTEOS}'.")
    ws = wb[HOJA_CONTEOS]

    # Mapear columna -> nombre de cuenta usando la fila de encabezado.
    col_cuenta: dict[int, str] = {}
    col_fecha = col_tc = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(FILA_ENCABEZADO, c).value
        if not isinstance(val, str):
            continue
        v = val.strip().upper()
        if v == "FECHA":
            col_fecha = c
        elif v == "T/C":
            col_tc = c
        elif v in COLUMNA_A_CUENTA:
            col_cuenta[c] = COLUMNA_A_CUENTA[v]
    if col_fecha is None or col_tc is None:
        sys.exit("No se encontraron las columnas FECHA / T/C en CONTEOS.")

    fotos = []
    descartadas = []
    for r in range(FILA_ENCABEZADO + 1, ws.max_row + 1):
        fecha = ws.cell(r, col_fecha).value
        tc = ws.cell(r, col_tc).value
        # Solo filas con FECHA real y T/C numérico son fotos válidas.
        if not isinstance(fecha, dt.datetime) or not isinstance(tc, (int, float)):
            # Registrar si tenía algo (datos sucios) para el reporte.
            if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
                descartadas.append(r)
            continue
        balances = {}
        for c, nombre in col_cuenta.items():
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)):
                balances[nombre] = float(val)
        fotos.append({"fila": r, "fecha": fecha.date(), "tc": float(tc), "balances": balances})
    return fotos, descartadas


def calcular_total_bob(balances: dict[str, float], tc: float, cuentas: dict[str, dict]) -> float:
    total = 0.0
    for nombre, amount in balances.items():
        cta = cuentas.get(nombre)
        if not cta:
            continue
        en_bob = amount if cta["currency"] == "BOB" else amount * tc
        total += -en_bob if cta["is_liability"] else en_bob
    return round(total, 2)


def main():
    ap = argparse.ArgumentParser(description="Migra CONTEOS del Excel a Supabase.")
    ap.add_argument("--excel", required=True, help="Ruta a My_Money_v5.0.xlsx")
    ap.add_argument("--commit", action="store_true", help="Escribe en la DB (por defecto: dry-run)")
    args = ap.parse_args()

    cargar_env()
    fotos, descartadas = leer_conteos(args.excel)

    print("\n=== RESUMEN DE MIGRACIÓN (CONTEOS -> patrimonio) ===")
    print(f"  Fotos válidas encontradas: {len(fotos)}")
    print(f"  Filas descartadas (datos sucios): {len(descartadas)} -> filas {descartadas}")

    conn = conectar()
    conn.autocommit = False
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        user_id = obtener_user_id(cur)
        cuentas = obtener_cuentas(cur, user_id)

        # Fechas ya existentes (idempotencia).
        cur.execute(
            "select snapshot_date from net_worth_snapshots where user_id = %s;", (user_id,)
        )
        existentes = {r["snapshot_date"] for r in cur.fetchall()}

        nuevas = [f for f in fotos if f["fecha"] not in existentes]
        saltadas = [f for f in fotos if f["fecha"] in existentes]

        print(f"  Ya existen (se saltan): {len(saltadas)}")
        print(f"  A insertar: {len(nuevas)}")
        print("\n  Detalle de fotos a insertar:")
        for f in nuevas:
            total_bob = calcular_total_bob(f["balances"], f["tc"], cuentas)
            total_usd = round(total_bob / f["tc"], 2) if f["tc"] else 0.0
            print(
                f"    - {f['fecha']}  T/C={f['tc']:.2f}  "
                f"cuentas={len(f['balances'])}  total_bob={total_bob:,.2f}  total_usd={total_usd:,.2f}"
            )

        if not args.commit:
            print("\n[DRY-RUN] No se escribió nada. Repite con --commit para cargar.\n")
            conn.rollback()
            return

        insertadas = 0
        for f in nuevas:
            total_bob = calcular_total_bob(f["balances"], f["tc"], cuentas)
            total_usd = round(total_bob / f["tc"], 2) if f["tc"] else 0.0
            cur.execute(
                """insert into net_worth_snapshots
                   (user_id, snapshot_date, exchange_rate, total_bob, total_usd)
                   values (%s, %s, %s, %s, %s) returning id;""",
                (user_id, f["fecha"], f["tc"], total_bob, total_usd),
            )
            snap_id = cur.fetchone()["id"]
            for nombre, amount in f["balances"].items():
                cta = cuentas.get(nombre)
                if not cta:
                    continue
                cur.execute(
                    """insert into net_worth_balances
                       (user_id, snapshot_id, account_id, amount)
                       values (%s, %s, %s, %s);""",
                    (user_id, snap_id, cta["id"], amount),
                )
            insertadas += 1

        conn.commit()
        print(f"\n[OK] {insertadas} fotos insertadas.\n")


if __name__ == "__main__":
    main()
